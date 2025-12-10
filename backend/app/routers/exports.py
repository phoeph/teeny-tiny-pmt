from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
from typing import Optional, List, Tuple, Dict
import io
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func
from app.database import get_db
from app.models import WorkItem, Project, User
from app.dependencies.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["exports"])

def _parse_ids(s: Optional[str]) -> List[int]:
    if not s:
        return []
    ids: List[int] = []
    for part in str(s).split(','):
        part = part.strip()
        if not part:
            continue
        try:
            ids.append(int(part))
        except Exception:
            continue
    return ids

def _to_date_str(d) -> str:
    try:
        return d.isoformat()
    except Exception:
        return ""

@router.get("/jobs-tasks.xlsx")
async def export_jobs_tasks(
    dimension: str = Query("project"),
    project_ids: Optional[str] = Query(None),
    assignee_ids: Optional[str] = Query(None),
    role: str = Query("both"),
    status: Optional[str] = Query(None),
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
    include_deleted: bool = Query(False),
    archived: bool = Query(False),
    tz: Optional[str] = Query(None),
    locale: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p_ids = _parse_ids(project_ids)
    a_ids = _parse_ids(assignee_ids)
    statuses = [s.strip() for s in (status or '').split(',') if s.strip()]

    wi_filters = []
    if p_ids:
        wi_filters.append(WorkItem.project_id.in_(p_ids))
    if a_ids:
        if role == 'owner':
            wi_filters.append(or_(Project.owner_id.in_(a_ids)))
        elif role == 'assignee':
            wi_filters.append(or_(WorkItem.assignee_id.in_(a_ids)))
        else:
            wi_filters.append(or_(WorkItem.assignee_id.in_(a_ids), Project.owner_id.in_(a_ids)))
    if statuses:
        wi_filters.append(WorkItem.status.in_(statuses))
    if not include_deleted:
        wi_filters.append(WorkItem.deleted_at.is_(None))
    if archived:
        wi_filters.append(Project.status == 'archived')
    else:
        wi_filters.append(Project.status != 'archived')

    # 时间范围交集：planned范围与查询(start/end)有重叠
    from datetime import date
    start_d = None
    end_d = None
    try:
        start_d = date.fromisoformat(start) if start else None
        end_d = date.fromisoformat(end) if end else None
    except Exception:
        start_d = None
        end_d = None
    if start_d:
        wi_filters.append(or_(WorkItem.planned_end_date.is_(None), WorkItem.planned_end_date >= start_d))
    if end_d:
        wi_filters.append(or_(WorkItem.planned_start_date.is_(None), WorkItem.planned_start_date <= end_d))

    base_stmt = select(WorkItem, Project, User).join(Project, WorkItem.project_id == Project.id).join(User, WorkItem.assignee_id == User.id, isouter=True)
    if wi_filters:
        base_stmt = base_stmt.where(and_(*wi_filters))

    jobs_stmt = base_stmt.where(WorkItem.kind == 'JOB')
    tasks_stmt = base_stmt.where(WorkItem.kind == 'TASK')

    jobs_res = await db.execute(jobs_stmt)
    tasks_res = await db.execute(tasks_stmt)
    jobs_rows = jobs_res.all()
    tasks_rows = tasks_res.all()

    # 预取父JOB code映射
    parent_map: Dict[int, str] = {}
    for wi, prj, usr in jobs_rows:
        parent_map[wi.id] = wi.code

    wb = Workbook()
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"
    ws_jobs.append(["ID","标题","项目","Owner","Assignee","状态","开始时间","结束时间","预计工时","记录工时","进度"]) 
    for wi, prj, usr in jobs_rows:
        owner_name = ''
        try:
            owner_stmt = await db.execute(select(User).where(User.id == prj.owner_id))
            owner_user = owner_stmt.scalar_one_or_none()
            owner_name = (owner_user.full_name or owner_user.username or owner_user.email_prefix) if owner_user else ''
        except Exception:
            owner_name = ''
        assignee_name = ''
        if usr:
            assignee_name = usr.full_name or usr.username or usr.email_prefix or ''
        pct = ''
        try:
            if wi.estimated_hours and wi.actual_hours is not None and wi.estimated_hours > 0:
                pct = f"{round((wi.actual_hours/wi.estimated_hours)*100)}%"
        except Exception:
            pct = ''
        ws_jobs.append([
            wi.code,
            wi.title,
            prj.code if hasattr(prj, 'code') else (prj.name or ''),
            owner_name,
            assignee_name,
            wi.status,
            _to_date_str(wi.planned_start_date or wi.start_date),
            _to_date_str(wi.planned_end_date or wi.end_date),
            wi.estimated_hours or '',
            wi.actual_hours or '',
            pct,
        ])

    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(["ID","标题","项目","Owner","Assignee","状态","开始时间","结束时间","预计工时","记录工时","ParentID"]) 
    for wi, prj, usr in tasks_rows:
        owner_name = ''
        try:
            owner_stmt = await db.execute(select(User).where(User.id == prj.owner_id))
            owner_user = owner_stmt.scalar_one_or_none()
            owner_name = (owner_user.full_name or owner_user.username or owner_user.email_prefix) if owner_user else ''
        except Exception:
            owner_name = ''
        assignee_name = ''
        if usr:
            assignee_name = usr.full_name or usr.username or usr.email_prefix or ''
        parent_code = ''
        try:
            if wi.parent_id:
                parent_code = parent_map.get(wi.parent_id, '')
                if not parent_code:
                    p_stmt = await db.execute(select(WorkItem.code).where(WorkItem.id == wi.parent_id))
                    parent_code = p_stmt.scalar_one_or_none() or ''
        except Exception:
            parent_code = ''
        ws_tasks.append([
            wi.code,
            wi.title,
            prj.code if hasattr(prj, 'code') else (prj.name or ''),
            owner_name,
            assignee_name,
            wi.status,
            _to_date_str(wi.planned_start_date or wi.start_date),
            _to_date_str(wi.planned_end_date or wi.end_date),
            wi.estimated_hours or '',
            wi.actual_hours or '',
            parent_code,
        ])

    # Summary：按项目与负责人聚合
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["项目","负责人","预计工时","记录工时","偏差"]) 
    # 聚合数据来源：jobs + tasks
    agg: Dict[Tuple[str,str], Tuple[float,float]] = {}
    def add_agg(proj_name: str, assignee_name: str, est: float, act: float):
        key = (proj_name or '', assignee_name or '')
        cur = agg.get(key, (0.0, 0.0))
        agg[key] = (cur[0] + (est or 0.0), cur[1] + (act or 0.0))
    for w, p, u in jobs_rows:
        pname = p.code if hasattr(p, 'code') else (p.name or '')
        aname = (u.full_name or u.username or u.email_prefix) if u else ''
        add_agg(pname, aname, w.estimated_hours or 0.0, w.actual_hours or 0.0)
    for w, p, u in tasks_rows:
        pname = p.code if hasattr(p, 'code') else (p.name or '')
        aname = (u.full_name or u.username or u.email_prefix) if u else ''
        add_agg(pname, aname, w.estimated_hours or 0.0, w.actual_hours or 0.0)
    for (pname, aname), (est, act) in agg.items():
        ws_sum.append([pname, aname, est, act, (est - act)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=jobs-tasks.xlsx"}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
