"""
标签树路由：从 .docs/tech/菜单级联关系.xlsx 解析为树形JSON
"""
from fastapi import APIRouter
from pathlib import Path
from typing import List, Dict, Any

router = APIRouter(prefix="/api/labels", tags=["标签树"])

_CACHE: Dict[str, Any] | None = None


def _excel_path() -> Path:
    # 解析仓库根目录下 .docs/tech/菜单级联关系.xlsx
    base = Path(__file__).resolve().parents[3]
    return base / ".docs" / "tech" / "菜单级联关系.xlsx"


def _build_tree(rows: List[List[str]]) -> List[Dict[str, Any]]:
    # 逐行构建多级树：每行对应一条路径，如 [一级, 二级, 三级, ...]
    root: Dict[str, Any] = {"children": {}}  # 使用字典构造，最后再转为数组

    def ensure_child(node: Dict[str, Any], name: str) -> Dict[str, Any]:
        ch = node.setdefault("children", {})
        if name not in ch:
            ch[name] = {"name": name, "children": {}}
        return ch[name]

    for r in rows:
        path = [s.strip() for s in r if s and str(s).strip()]
        if not path:
            continue
        cur = root
        for i, name in enumerate(path):
            cur = ensure_child(cur, name)
            if i == len(path) - 1:
                # 叶子：记录完整路径与唯一ID（用路径字符串）
                cur["path"] = "/".join(path)
                cur["id"] = cur["path"]
    # 将字典形式 children 转数组形式
    def dict_to_list(n: Dict[str, Any]) -> List[Dict[str, Any]]:
        arr: List[Dict[str, Any]] = []
        for _, v in (n.get("children") or {}).items():
            node = {"name": v["name"]}
            if "id" in v:
                node["id"] = v["id"]
                node["path"] = v["path"]
            children = dict_to_list(v)
            if children:
                node["children"] = children
            arr.append(node)
        return arr
    return dict_to_list(root)


def _parse_excel() -> List[Dict[str, Any]]:
    from openpyxl import load_workbook
    p = _excel_path()
    if not p.exists():
        return []
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        # 每行多个层级（按列），过滤空白
        vals = []
        for c in row:
            if c is None:
                vals.append("")
            else:
                vals.append(str(c))
        rows.append(vals)
    # 去掉表头（如存在），策略：若首行包含“一级/二级/三级”等字样，则跳过
    if rows:
        header = "/".join([s.strip() for s in rows[0] if s and s.strip()])
        if any(k in header for k in ["一级", "二级", "三级", "四级", "五级", "Level", "层级"]):
            rows = rows[1:]
    return _build_tree(rows)


@router.get("/tree")
async def get_label_tree():
    global _CACHE
    if _CACHE is None:
        _CACHE = {"items": _parse_excel()}
    return _CACHE

